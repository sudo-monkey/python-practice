import openpyxl

load_workbook = openpyxl.load_workbook('testfile.xlsx')
wb = openpyxl.load_workbook(filename='testfile.xlsx')

ws1 = wb.worksheets[0]

ws1['A1'] = 1
ws1.cell(row=2, column=2).value = 2
ws1.cell(coordinate="C3").value = 3
